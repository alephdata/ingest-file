import logging
from followthemoney import model
from openpyxl import load_workbook
from xml.etree.ElementTree import ParseError

from ingestors.ingestor import Ingestor
from ingestors.support.table import TableSupport
from ingestors.support.ooxml import OOXMLSupport
from ingestors.exc import ProcessingException

log = logging.getLogger(__name__)


class ExcelXMLIngestor(Ingestor, TableSupport, OOXMLSupport):
    MIME_TYPES = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # noqa
        "application/vnd.openxmlformats-officedocument.spreadsheetml.template",  # noqa
        "application/vnd.ms-excel.sheet.macroenabled.12",
        "application/vnd.ms-excel.sheet.binary.macroenabled.12",
        "application/vnd.ms-excel.template.macroenabled.12",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    ]
    EXTENSIONS = ["xlsx", "xlsm", "xltx", "xltm"]
    SCORE = 7

    def generate_rows(self, sheet):
        for row in sheet.rows:
            try:
                yield [c.value for c in row]
            except (ValueError, OverflowError, ParseError) as ve:
                log.warning("Failed to read Excel row: %s", ve)

    def ingest(self, file_path, entity):
        entity.schema = model.get("Workbook")
        self.ooxml_extract_metadata(file_path, entity)
        try:
            book = load_workbook(file_path, read_only=True)
        except Exception as err:
            raise ProcessingException("Invalid Excel file: %s" % err) from err

        try:
            for name in book.sheetnames:
                sheet = book[name]
                if not hasattr(sheet, "rows"):
                    log.warning("Cannot parse chart sheet: %s", name)
                    continue
                table = self.manager.make_entity("Table", parent=entity)
                table.make_id(entity.id, name)
                table.set("title", name)
                # Emit a partial table fragment with parent reference and name
                # early, so that we don't have orphan fragments in case of an error
                # in the middle of processing.
                # See https://github.com/alephdata/ingest-file/issues/171
                self.manager.emit_entity(table, fragment="initial")
                log.debug("Sheet: %s", name)
                self.emit_row_tuples(table, self.generate_rows(sheet))
                if table.has("csvHash"):
                    self.manager.emit_entity(table)
        except Exception as err:
            raise ProcessingException("Cannot read Excel file: %s" % err) from err
        finally:
            book.close()

    @classmethod
    def match(cls, file_path, entity):
        score = super(ExcelXMLIngestor, cls).match(file_path, entity)
        if score > 0 and not cls.inspect_ooxml_manifest(file_path):
            return -1
        return score
